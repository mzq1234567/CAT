"""
Client-facing cost-assessment report (PDF) + workbook (Excel).

`generate_pdf` renders the branded TechPlus Talent template: cover page, table of contents,
executive summary with 3-year spend projections, Purpose, Methodology, Evaluation Summary
(environment details), and Cost Optimization broken out by pillar (Compute / Storage / Network) with
per-optimization tables (Reserved Instances 1yr/3yr, Azure Hybrid Benefit, Savings Plan, right-sizing,
orphaned disks, public IPs). Numbers come straight from the findings — which are now grounded in
actual cost — so the report never shows savings exceeding spend.

`generate_excel` remains the detailed data workbook for analysts.
"""
from __future__ import annotations

import io
import os
from collections import OrderedDict
from typing import Dict, List, Optional

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    HRFlowable, Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from reportlab.platypus.tableofcontents import TableOfContents
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.legends import Legend

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ── Brand palette (TPT client template) ──────────────────────────────────────────
TEAL = colors.HexColor("#0E7C86")      # section headings
HEAD_BLUE = colors.HexColor("#2E75B6")  # table header fill
SAVINGS = colors.HexColor("#2E7D32")    # savings (green)
SPEND = colors.HexColor("#2E75B6")      # current spend (blue)
AFTER = colors.HexColor("#ED7D31")      # optimized spend (orange)
ROW_ALT = colors.HexColor("#F2F6FA")
INK = colors.HexColor("#1A2B3C")
MUTED = colors.HexColor("#5B6B7B")

_LOGO_PATH = os.path.join(os.path.dirname(__file__), "..", "assets", "tpt_logo.png")

# Helvetica-safe currency symbols (₹/native glyphs don't render in the base PDF font, so INR → "Rs ";
# the $-currencies are disambiguated with a country prefix).
_CURRENCY_SYMBOLS = {"USD": "$", "CAD": "CA$", "AUD": "A$", "GBP": "£", "EUR": "€", "INR": "Rs "}
_CUR = {"code": "USD", "symbol": "$"}


def _set_currency(code: Optional[str]) -> None:
    code = (code or "USD").upper()
    _CUR["code"] = code
    _CUR["symbol"] = _CURRENCY_SYMBOLS.get(code, code + " ")


def _usd(v: Optional[float], dec: int = 2) -> str:
    return f"{_CUR['symbol']}{(v or 0):,.{dec}f}"


def _logo() -> Optional[str]:
    return _LOGO_PATH if os.path.exists(_LOGO_PATH) else None


# ── Findings organisation ─────────────────────────────────────────────────────────

# category → pillar. Everything database-ish folds into Compute (commitment-style), matching the
# 3-pillar client template (Compute / Storage / Network).
_PILLAR = {
    "ri_vm": "Compute", "savings_plan_vm": "Compute", "windows_ahb": "Compute",
    "oversized_vms": "Compute", "idle_vms": "Compute", "deallocated_vms": "Compute",
    "vm_rightsizing": "Compute", "idle_app_service_plans": "Compute",
    "app_service_reserved_capacity": "Compute", "app_service_plan_rightsizing": "Compute",
    "paused_sql_databases": "Compute", "stopped_sql_managed_instances": "Compute",
    "sql_db_reserved_capacity": "Compute", "sql_mi_reserved_capacity": "Compute",
    "cosmos_reserved_capacity": "Compute", "mysql_reserved_capacity": "Compute",
    "unattached_managed_disks": "Storage", "orphaned_snapshots": "Storage",
    "managed_disk_reserved_capacity": "Storage", "disk_rightsizing": "Storage",
    "backup_redundancy": "Storage", "azure_files_reserved_capacity": "Storage",
    "orphaned_public_ips": "Network", "empty_load_balancers": "Network",
    "idle_nat_gateways": "Network", "bastion_hosts": "Network",
}


def _by_category(findings: List) -> "OrderedDict[str, List]":
    d: "OrderedDict[str, List]" = OrderedDict()
    for f in findings:
        d.setdefault(f.category, []).append(f)
    return d


def _one(cat: Dict, key: str):
    """First finding for an aggregated category (ri_vm / savings_plan_vm / windows_ahb), or None."""
    lst = cat.get(key)
    return lst[0] if lst else None


def _pillar_totals(findings: List) -> Dict[str, float]:
    totals = {"Compute": 0.0, "Storage": 0.0, "Network": 0.0, "Other": 0.0}
    for f in findings:
        pillar = _PILLAR.get(f.category, "Other")
        totals[pillar] += f.estimated_savings_annual or 0.0
    return totals


# ── Charts ─────────────────────────────────────────────────────────────────────────

_GROWTH = OrderedDict([
    ("Fixed Consumption Spend Trajectory", 0.0),
    ("Conservative Growth ACR", 0.07),
    ("Linear Growth ACR", 0.15),
    ("Civo 2024 SMB Report (25%) ACR", 0.25),
])


def _projection(spend: float, savings: float, g: float):
    """Cumulative 3-year (ACR, ACR-after-savings, savings) with annual growth rate `g`."""
    rate = (savings / spend) if spend else 0.0
    acr, after, sav = [], [], []
    c_acr = c_after = c_sav = 0.0
    for y in range(3):
        annual = spend * ((1 + g) ** y)
        s = annual * rate
        c_acr += annual
        c_sav += s
        c_after += annual - s
        acr.append(round(c_acr)); after.append(round(c_after)); sav.append(round(c_sav))
    return acr, after, sav


def _bar_chart(series, cats, colours, width=470, height=210, legend_names=None):
    d = Drawing(width, height)
    bc = VerticalBarChart()
    bc.x, bc.y = 45, 45
    bc.width, bc.height = width - 90, height - 80
    bc.data = series
    bc.categoryAxis.categoryNames = cats
    bc.categoryAxis.labels.fontSize = 8
    bc.valueAxis.valueMin = 0
    bc.valueAxis.labels.fontSize = 7
    bc.barSpacing = 1
    bc.groupSpacing = 12
    for i, col in enumerate(colours):
        bc.bars[i].fillColor = col
        bc.bars[i].strokeColor = colors.white
    # $ value label on top of every bar (like the client template).
    bc.barLabelFormat = f"{_CUR['symbol']}%0.0f"
    bc.barLabels.fontName = "Helvetica"
    bc.barLabels.fontSize = 6
    bc.barLabels.dy = 4
    bc.barLabels.fillColor = INK
    d.add(bc)
    if legend_names:
        lg = Legend()
        lg.x, lg.y = 45, 16
        lg.deltax = 120
        lg.dxTextSpace = 5
        lg.columnMaximum = 1
        lg.alignment = "right"
        lg.fontSize = 8
        lg.colorNamePairs = list(zip(colours, legend_names))
        d.add(lg)
    return d


def _savings_options_chart(cat_finding):
    """Compute-pillar options bar: RI 1yr / RI 3yr / AHB / Savings Plan annual savings."""
    labels, values, cols = [], [], []
    ri = _one(cat_finding, "ri_vm")
    if ri:
        det = ri.details or {}
        one = (det.get("total_1yr_monthly") or 0) * 12
        three = (det.get("total_3yr_monthly") or 0) * 12
        if one:
            labels.append("RI (1 Year)"); values.append(round(one)); cols.append(SAVINGS)
        if three:
            labels.append("RI (3 Year)"); values.append(round(three)); cols.append(AFTER)
    sp = _one(cat_finding, "savings_plan_vm")
    if sp:
        labels.append("Savings Plan"); values.append(round(sp.estimated_savings_annual)); cols.append(SPEND)
    ahb = _one(cat_finding, "windows_ahb")
    if ahb:
        labels.append("Azure Hybrid Benefit"); values.append(round(ahb.estimated_savings_annual)); cols.append(TEAL)
    if not values:
        return None
    d = Drawing(470, 200)
    bc = VerticalBarChart()
    bc.x, bc.y = 45, 35
    bc.width, bc.height = 380, 140
    bc.data = [values]
    bc.categoryAxis.categoryNames = labels
    bc.categoryAxis.labels.fontSize = 8
    bc.valueAxis.valueMin = 0
    bc.valueAxis.labels.fontSize = 7
    bc.barWidth = 18
    for i in range(len(values)):
        bc.bars[(0, i)].fillColor = cols[i]
    bc.barLabelFormat = f"{_CUR['symbol']}%0.0f"
    bc.barLabels.fontName = "Helvetica-Bold"
    bc.barLabels.fontSize = 7
    bc.barLabels.dy = 4
    bc.barLabels.fillColor = INK
    d.add(bc)
    return d


# ── PDF document (TOC-aware) ────────────────────────────────────────────────────────

class _TPTDoc(SimpleDocTemplate):
    """SimpleDocTemplate that notifies the TableOfContents of H1/H2 headings for page numbers."""

    def afterFlowable(self, flowable):
        if not isinstance(flowable, Paragraph):
            return
        name = flowable.style.name
        if name == "TOCH1":
            self.notify("TOCEntry", (0, flowable.getPlainText(), self.page))
        elif name == "TOCH2":
            self.notify("TOCEntry", (1, flowable.getPlainText(), self.page))


def _cover(canvas, doc):
    """Full-page dark cover drawn under the page-1 flowables."""
    canvas.saveState()
    w, h = A4
    canvas.setFillColor(colors.HexColor("#0B1622"))
    canvas.rect(0, 0, w, h, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#123047"))
    canvas.rect(0, 0, w, 2.4 * inch, fill=1, stroke=0)  # subtle base band
    canvas.restoreState()


def _decorate(canvas, doc):
    """Header logo + footer page number on content pages."""
    canvas.saveState()
    w, h = A4
    logo = _logo()
    if logo:
        try:
            canvas.drawImage(logo, w - 1.7 * inch, h - 0.75 * inch, width=1.2 * inch,
                             height=0.42 * inch, preserveAspectRatio=True, mask="auto")
        except Exception:  # noqa: BLE001
            pass
    canvas.setStrokeColor(colors.HexColor("#DDE5EE"))
    canvas.setLineWidth(0.5)
    canvas.line(0.7 * inch, 0.6 * inch, w - 0.7 * inch, 0.6 * inch)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(MUTED)
    canvas.drawRightString(w - 0.7 * inch, 0.42 * inch, f"{doc.page} | P a g e")
    canvas.restoreState()


def _styles():
    s = getSampleStyleSheet()
    return {
        "cover_title": ParagraphStyle("cover_title", fontName="Helvetica-Bold", fontSize=40,
                                      textColor=colors.white, leading=46),
        "cover_meta": ParagraphStyle("cover_meta", fontName="Helvetica-Bold", fontSize=16,
                                     textColor=colors.white, leading=22),
        "h1": ParagraphStyle("TOCH1", fontName="Helvetica-Bold", fontSize=17, textColor=TEAL,
                             spaceBefore=6, spaceAfter=8),
        "h2": ParagraphStyle("TOCH2", fontName="Helvetica-Bold", fontSize=13, textColor=TEAL,
                             spaceBefore=8, spaceAfter=5),
        "h3": ParagraphStyle("h3", fontName="Helvetica-Bold", fontSize=10.5, textColor=INK,
                             spaceBefore=6, spaceAfter=3),
        "body": ParagraphStyle("body", parent=s["Normal"], fontSize=9.5, textColor=INK, leading=14,
                               spaceAfter=4),
        "small": ParagraphStyle("small", parent=s["Normal"], fontSize=8, textColor=MUTED, leading=11),
        "cell": ParagraphStyle("cell", parent=s["Normal"], fontSize=8, textColor=INK, leading=10),
        "cellh": ParagraphStyle("cellh", parent=s["Normal"], fontSize=8, textColor=colors.white,
                                fontName="Helvetica-Bold", leading=10),
        "toch": ParagraphStyle("toch", fontName="Helvetica-Bold", fontSize=16, textColor=TEAL,
                               spaceAfter=10, alignment=TA_CENTER),
    }


def _table(headers, rows, col_widths, st, align_right_from=None):
    """A branded data table: blue header, alternating rows, right-aligned money columns."""
    data = [[Paragraph(h, st["cellh"]) for h in headers]]
    for r in rows:
        data.append([c if hasattr(c, "wrap") else Paragraph(str(c), st["cell"]) for c in r])
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), HEAD_BLUE),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6DEE8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ]
    if align_right_from is not None:
        style.append(("ALIGN", (align_right_from, 1), (-1, -1), "RIGHT"))
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), ROW_ALT))
    t.setStyle(TableStyle(style))
    return t


def _hero(assessment, st):
    """Executive-summary savings hero (3-year / yearly / monthly)."""
    annual = assessment.total_savings_annual or 0.0
    monthly = assessment.total_savings_monthly or 0.0
    cells = [
        [Paragraph("Potential 3-Year Savings", st["cellh"]),
         Paragraph("Potential Yearly Savings", st["cellh"]),
         Paragraph("Potential Monthly Savings", st["cellh"])],
        [Paragraph(f"<b>{_usd(annual * 3)}</b>", ParagraphStyle("hb", parent=st["body"], fontSize=18,
                                                                textColor=SAVINGS, alignment=TA_CENTER)),
         Paragraph(f"<b>{_usd(annual)}</b>", ParagraphStyle("hy", parent=st["body"], fontSize=18,
                                                            textColor=SAVINGS, alignment=TA_CENTER)),
         Paragraph(f"<b>{_usd(monthly)}</b>", ParagraphStyle("hm", parent=st["body"], fontSize=18,
                                                             textColor=SAVINGS, alignment=TA_CENTER))],
    ]
    t = Table(cells, colWidths=[2.3 * inch] * 3)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HEAD_BLUE),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EAF4EC")),
        ("TOPPADDING", (0, 1), (-1, 1), 8), ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
    ]))
    return t


# ── Pillar detail builders ──────────────────────────────────────────────────────────

def _compute_section(cat: Dict, st) -> List:
    out: List = []
    chart = _savings_options_chart(cat)
    if chart:
        out.append(Paragraph("Potential Annual Savings by Optimization Option", st["h3"]))
        out.append(chart)
        out.append(Spacer(1, 6))

    ri = _one(cat, "ri_vm")
    if ri and (ri.details or {}).get("reservation_items"):
        items = ri.details["reservation_items"]
        out.append(Paragraph("Reserved Instances (1 Year)", st["h3"]))
        out.append(Paragraph(
            "Virtual machines expected to run long-term. Purchasing a 1-year Reserved Instance for "
            "these achieves the savings below versus pay-as-you-go.", st["small"]))
        out.append(Spacer(1, 3))
        rows = [[i.get("name") or i.get("sku"), i.get("sku"), i.get("region") or "—",
                 i.get("quantity") or "—",
                 _usd((i.get("monthly_ondemand") or 0) * 12), _usd((i.get("monthly_savings") or 0) * 12)]
                for i in items]
        out.append(_table(["Name", "SKU", "Region", "Qty", "Current Annual", "Annual Savings (1 Yr)"],
                          rows, [1.5 * inch, 1.3 * inch, 0.9 * inch, 0.5 * inch, 1.1 * inch, 1.4 * inch],
                          st, align_right_from=4))
        out.append(Spacer(1, 8))
        three = [i for i in items if i.get("monthly_savings_3yr")]
        if three:
            out.append(Paragraph("Reserved Instances (3 Year)", st["h3"]))
            rows3 = [[i.get("name") or i.get("sku"), i.get("sku"), i.get("region") or "—",
                      i.get("quantity") or "—", _usd((i.get("monthly_savings_3yr") or 0) * 12)]
                     for i in three]
            out.append(_table(["Name", "SKU", "Region", "Qty", "Annual Savings (3 Yr)"], rows3,
                              [1.7 * inch, 1.4 * inch, 1.0 * inch, 0.6 * inch, 1.5 * inch],
                              st, align_right_from=4))
            out.append(Spacer(1, 8))

    sp = _one(cat, "savings_plan_vm")
    if sp and (sp.details or {}).get("reservation_items"):
        items = sp.details["reservation_items"]
        out.append(Paragraph("Savings Plan (Compute)", st["h3"]))
        rows = [[i.get("name") or i.get("sku"), i.get("sku"), i.get("region") or "—",
                 _usd((i.get("monthly_savings") or 0) * 12)] for i in items]
        out.append(_table(["Name", "SKU", "Region", "Annual Savings (1 Yr)"], rows,
                          [2.0 * inch, 1.6 * inch, 1.2 * inch, 1.5 * inch], st, align_right_from=3))
        out.append(Spacer(1, 8))

    ahb = _one(cat, "windows_ahb")
    if ahb and (ahb.details or {}).get("eligible_vms"):
        vms = ahb.details["eligible_vms"]
        out.append(Paragraph("Azure Hybrid Benefit", st["h3"]))
        out.append(Paragraph(
            "Windows VMs eligible to apply Azure Hybrid Benefit (requires eligible Windows Server "
            "licences with Software Assurance). Savings shown are the Windows licence portion of each "
            "VM's actual cost.", st["small"]))
        out.append(Spacer(1, 3))
        rows = [[v.get("name"), v.get("sku"), _usd((v.get("monthly_savings") or 0) * 12)] for v in vms]
        out.append(_table(["Name", "Size", "Annual Savings"], rows,
                          [2.6 * inch, 1.8 * inch, 1.6 * inch], st, align_right_from=2))
        out.append(Spacer(1, 8))

    oversized = cat.get("oversized_vms") or []
    if oversized:
        out.append(Paragraph("Right-Sizing (under-utilised VMs)", st["h3"]))
        rows = []
        for f in oversized:
            d = f.details or {}
            rows.append([f.resource_name, d.get("current_sku") or "—", d.get("recommended_sku") or "—",
                         _usd(f.estimated_savings_annual)])
        out.append(_table(["Name", "Current Size", "New Size", "Annual Savings"], rows,
                          [1.9 * inch, 1.6 * inch, 1.6 * inch, 1.3 * inch], st, align_right_from=3))
        out.append(Spacer(1, 8))

    idle = (cat.get("idle_vms") or []) + (cat.get("deallocated_vms") or [])
    if idle:
        out.append(Paragraph("Idle / Deallocated VMs", st["h3"]))
        rows = [[f.resource_name, f.resource_type.split("/")[-1] if f.resource_type else "VM",
                 _usd(f.estimated_savings_annual)] for f in idle]
        out.append(_table(["Name", "Type", "Annual Savings"], rows,
                          [2.8 * inch, 1.6 * inch, 1.4 * inch], st, align_right_from=2))
        out.append(Spacer(1, 8))
    return out


def _storage_section(cat: Dict, st) -> List:
    out: List = []
    disks = cat.get("unattached_managed_disks") or []
    if disks:
        out.append(Paragraph("Orphaned (Unattached) Disks", st["h3"]))
        out.append(Paragraph(
            "Managed disks not attached to any VM. Validate they are not needed for backup/DR, then "
            "delete to stop the storage charge.", st["small"]))
        out.append(Spacer(1, 3))
        rows = []
        for f in disks:
            d = f.details or {}
            rows.append([f.resource_name, d.get("skuName") or "—", _usd(f.estimated_savings_annual)])
        out.append(_table(["Name", "Storage Type", "Annual Savings"], rows,
                          [3.6 * inch, 1.4 * inch, 1.4 * inch], st, align_right_from=2))
        out.append(Spacer(1, 8))
    for cat_key, title in (("orphaned_snapshots", "Orphaned Snapshots"),
                           ("backup_redundancy", "Backup Redundancy (GRS → LRS)"),
                           ("managed_disk_reserved_capacity", "Managed Disk Reserved Capacity")):
        fs = cat.get(cat_key) or []
        if fs:
            out.append(Paragraph(title, st["h3"]))
            rows = [[f.resource_name or "—", _usd(f.estimated_savings_annual)] for f in fs]
            out.append(_table(["Name", "Annual Savings"], rows, [5.0 * inch, 1.4 * inch], st,
                              align_right_from=1))
            out.append(Spacer(1, 8))
    return out


def _network_section(cat: Dict, st) -> List:
    out: List = []
    for cat_key, title, note in (
        ("orphaned_public_ips", "Orphaned Public IP Addresses",
         "Unattached public IPs — review and delete if no longer required."),
        ("empty_load_balancers", "Empty Load Balancers", ""),
        ("idle_nat_gateways", "Idle NAT Gateways", ""),
        ("bastion_hosts", "Azure Bastion (review)", ""),
    ):
        fs = cat.get(cat_key) or []
        if fs:
            out.append(Paragraph(title, st["h3"]))
            if note:
                out.append(Paragraph(note, st["small"]))
                out.append(Spacer(1, 3))
            rows = [[f.resource_name or "—", (f.details or {}).get("skuName") or "Standard",
                     _usd(f.estimated_savings_monthly)] for f in fs]
            out.append(_table(["Name", "SKU", "Cost / Month"], rows,
                              [3.6 * inch, 1.4 * inch, 1.4 * inch], st, align_right_from=2))
            out.append(Spacer(1, 8))
    return out


# ── PDF assembly ────────────────────────────────────────────────────────────────────

def generate_pdf(assessment, findings: List) -> bytes:
    _set_currency(getattr(assessment, "currency", None))
    st = _styles()
    buf = io.BytesIO()
    doc = _TPTDoc(buf, pagesize=A4, leftMargin=0.7 * inch, rightMargin=0.7 * inch,
                  topMargin=0.9 * inch, bottomMargin=0.8 * inch,
                  title="Cost Assessment", author="TechPlus Talent")

    active = [f for f in findings if not getattr(f, "dismissed", 0)]
    cat = _by_category(active)
    client = assessment.tenant_display_name or (
        list((assessment.subscription_names or {}).values() or [None])[0]) or "Your Organisation"
    when = (assessment.created_at.strftime("%B %d, %Y") if assessment.created_at else "")
    pillar_totals = _pillar_totals(active)
    spend_annual = assessment.current_annual_spend if assessment.cost_data_available else None

    story: List = []

    # ── Cover (page 1; dark background drawn by _cover) ─────────────────────────
    story.append(Spacer(1, 1.4 * inch))
    logo = _logo()
    if logo:
        try:
            story.append(Image(logo, width=2.2 * inch, height=0.8 * inch))
            story.append(Spacer(1, 0.5 * inch))
        except Exception:  # noqa: BLE001
            pass
    else:
        story.append(Paragraph("TECH<font color='#ED7D31'>PLUS</font>TALENT", st["cover_meta"]))
        story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph("Cost Assessment", st["cover_title"]))
    story.append(Spacer(1, 0.3 * inch))
    story.append(HRFlowable(width=1.6 * inch, thickness=2, color=colors.white, hAlign="LEFT"))
    story.append(Spacer(1, 0.25 * inch))
    if when:
        story.append(Paragraph(when, st["cover_meta"]))
    story.append(Paragraph(f"Prepared for: {client}", st["cover_meta"]))
    story.append(PageBreak())

    # ── Table of Contents ───────────────────────────────────────────────────────
    story.append(Paragraph("Table of Contents", st["toch"]))
    toc = TableOfContents()
    toc.levelStyles = [
        ParagraphStyle("toc1", fontName="Helvetica-Bold", fontSize=11, leading=18, textColor=INK),
        ParagraphStyle("toc2", fontName="Helvetica", fontSize=9.5, leading=15, leftIndent=16,
                       textColor=MUTED),
    ]
    story.append(toc)
    story.append(PageBreak())

    # ── 1. Executive Summary ────────────────────────────────────────────────────
    story.append(Paragraph("1.  Executive Summary", st["h1"]))
    story.append(Paragraph(
        f"As part of its cloud optimization strategy, <b>{client}</b> engaged <b>TechPlus Talent</b> "
        "to perform a comprehensive Cost Assessment of its Azure environment — evaluating spend "
        "patterns, analysing cost distribution across services, and identifying opportunities to "
        "optimize while maintaining operational effectiveness. Based on the Phase 1 assessment, the "
        "following savings have been identified:", st["body"]))
    story.append(Spacer(1, 8))
    story.append(_hero(assessment, st))
    story.append(Spacer(1, 10))
    if spend_annual:
        story.append(Paragraph(
            f"Current forecasted annual spend is <b>{_usd(spend_annual)}</b>; the recommendations "
            f"below identify up to <b>{_usd(assessment.total_savings_annual)}</b> in annual savings "
            f"(about {round((assessment.total_savings_annual or 0) / spend_annual * 100)}% of spend).",
            st["body"]))
        story.append(Spacer(1, 8))
        # 3-year projection scenarios.
        for title, g in _GROWTH.items():
            acr, after, sav = _projection(spend_annual, assessment.total_savings_annual or 0, g)
            story.append(Paragraph(f"{title} — 3 Years", st["h3"]))
            story.append(_bar_chart(
                [acr, after, sav], ["Year 1", "Year 2", "Year 3"], [SPEND, AFTER, SAVINGS],
                legend_names=["Projected spend (ACR)", "Spend after savings", "Cumulative savings"]))
            story.append(Spacer(1, 6))
    else:
        story.append(Paragraph(
            "Azure Cost Management billing data was not available for the signed-in account, so total "
            "spend and multi-year projections are omitted. The savings figures above are grounded in "
            "each resource's estimated cost.", st["small"]))
    story.append(PageBreak())

    # ── 2. Purpose ───────────────────────────────────────────────────────────────
    story.append(Paragraph("2.  Purpose", st["h1"]))
    for point in [
        f"Assess {client}'s Azure landscape with a focused review of compute, storage and network "
        "resources to understand key cost drivers and consumption trends.",
        "Validate that existing workloads adhere to Microsoft best practices for cost efficiency, "
        "data management, and long-term operational sustainability.",
        "Provide clear, actionable recommendations to optimize resource utilization, strengthen cost "
        "governance, and maximize cloud investment value.",
        "Support strategic decision-making by presenting transparent cost-saving projections, "
        "balancing a strong operational posture with financial efficiency.",
    ]:
        story.append(Paragraph(f"•  {point}", st["body"]))

    # ── 3. TPT Methodology ────────────────────────────────────────────────────────
    story.append(Paragraph("3.  TPT Methodology", st["h1"]))
    story.append(Paragraph("Phase 1: Quick Win Assessment", st["h2"]))
    story.append(Paragraph(
        "<b>Objective</b> — Assess the environment and provide recommendations that do not require "
        "architectural changes.<br/><b>Outcome</b> — Immediate cost savings through configuration "
        "tweaks, resource optimization, and policy adjustments.", st["body"]))
    story.append(Paragraph("Phase 2: Deep Dive Modernization", st["h2"]))
    story.append(Paragraph(
        "<b>Objective</b> — Comprehensive assessment of the environment, inside and out.<br/>"
        "<b>Outcome</b> — Architectural changes, modernization strategies, and advanced optimizations "
        "for further cost savings.", st["body"]))

    # ── 4. Evaluation Summary ─────────────────────────────────────────────────────
    story.append(Paragraph("4.  Evaluation Summary", st["h1"]))
    story.append(Paragraph("3 Pillars of Azure Cost Assessment: Compute · Storage · Network", st["body"]))
    story.append(Spacer(1, 4))
    subs = assessment.subscription_names or {}
    sub_str = "; ".join(f"{name or sid}" for sid, name in subs.items()) or \
        ", ".join(assessment.subscription_ids or [])
    major = assessment.major_resource_types or []
    major_str = ", ".join(m.get("type", "") for m in major) or "—"
    env_rows = [
        ["Tenant", f"{client}  ({assessment.tenant_id or '—'})"],
        ["Subscription(s)", sub_str or "—"],
        ["Total No. of resources", str(assessment.total_resources or "—")],
        ["Major Resource Types", major_str],
    ]
    if assessment.current_monthly_spend is not None:
        env_rows.append(["Last Month Consumption", _usd(assessment.current_monthly_spend)])
    et = Table([[Paragraph("Environment Details", st["cellh"]), ""]] +
               [[Paragraph(k, st["cell"]), Paragraph(str(v), st["cell"])] for k, v in env_rows],
               colWidths=[2.2 * inch, 4.4 * inch])
    et.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HEAD_BLUE), ("SPAN", (0, 0), (-1, 0)),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6DEE8")),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#EAF1F8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(et)
    story.append(Spacer(1, 8))
    story.append(Paragraph("Identified savings by pillar:", st["body"]))
    pillar_rows = [[p, _usd(pillar_totals.get(p, 0))] for p in ("Compute", "Storage", "Network")
                   if pillar_totals.get(p, 0) > 0]
    if pillar_totals.get("Other", 0) > 0:
        pillar_rows.append(["Other", _usd(pillar_totals["Other"])])
    if pillar_rows:
        story.append(_table(["Pillar", "Annual Savings"], pillar_rows,
                            [3.0 * inch, 1.6 * inch], st, align_right_from=1))
    story.append(PageBreak())

    # ── 5. Cost Optimization (by pillar) ──────────────────────────────────────────
    story.append(Paragraph("5.  Cost Optimization", st["h1"]))
    pillar_builders = [
        ("5.1  Compute", "Compute", _compute_section),
        ("5.2  Storage", "Storage", _storage_section),
        ("5.3  Network", "Network", _network_section),
    ]
    any_pillar = False
    for heading, pillar, builder in pillar_builders:
        section = builder(cat, st)
        if not section:
            continue
        any_pillar = True
        story.append(Paragraph(heading, st["h2"]))
        story.append(Paragraph(
            f"Identified cost-saving opportunities for {pillar.lower()} resources — "
            f"max yearly savings {_usd(pillar_totals.get(pillar, 0))}.", st["body"]))
        story.append(Spacer(1, 4))
        story.extend(section)
    if not any_pillar:
        story.append(Paragraph("No optimization opportunities were identified in this assessment.",
                               st["body"]))

    doc.multiBuild(story, onFirstPage=_cover, onLaterPages=_decorate)
    return buf.getvalue()


# ── Excel workbook (analyst detail) — unchanged shape ────────────────────────────────

def _confidence_label(conf: float) -> str:
    if conf >= 0.8:
        return "High"
    if conf >= 0.5:
        return "Medium"
    return "Low"


def _as_of(assessment) -> str:
    if getattr(assessment, "snapshot_at", None):
        return assessment.snapshot_at.strftime("%Y-%m-%d %H:%M UTC")
    return "N/A"


def _validation_counts(findings: List):
    counts = {"validated": 0, "needs_review": 0, "unvalidated": 0}
    for f in findings:
        status = (getattr(f, "validation_status", None) or "unvalidated").lower()
        counts[status] = counts.get(status, 0) + 1
    return counts


def generate_excel(assessment, findings: List) -> bytes:
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1565C0")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)

    ws_sum = wb.active
    ws_sum.title = "Summary"
    counts = _validation_counts(findings)
    summary_rows = [
        ("Assessment ID", assessment.id),
        ("Client (Tenant)", getattr(assessment, "tenant_display_name", None) or ""),
        ("User", assessment.user_email),
        ("Generated", assessment.created_at.strftime("%Y-%m-%d %H:%M UTC")),
        ("Data as of (snapshot)", _as_of(assessment)),
        ("Status", assessment.status.upper()),
        ("Subscriptions", len(assessment.subscription_ids)),
        ("Total Findings", assessment.findings_count),
        ("Current Annual Spend (USD)", getattr(assessment, "current_annual_spend", None)),
        ("Monthly Savings (USD)", assessment.total_savings_monthly),
        ("Annual Savings (USD)", assessment.total_savings_annual),
        ("Validated", counts["validated"]),
        ("Needs Review", counts["needs_review"]),
        ("Unvalidated", counts["unvalidated"]),
    ]
    for r, (label, value) in enumerate(summary_rows, 1):
        ws_sum.cell(r, 1, label).font = Font(bold=True)
        ws_sum.cell(r, 2, value)
    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 40

    ws = wb.create_sheet("Findings")
    headers = [
        "Category", "Resource Name", "Resource Group", "Subscription ID", "Resource Type",
        "Monthly Savings (USD)", "Annual Savings (USD)", "Severity", "Confidence",
        "Confidence Level", "Validation", "Variance %", "Actual Monthly Cost (USD)",
        "Advisor Rec ID", "Description", "Recommendation",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    sev_fills = {
        "critical": PatternFill("solid", fgColor="EF9A9A"),
        "high": PatternFill("solid", fgColor="FFCDD2"),
        "medium": PatternFill("solid", fgColor="FFF9C4"),
        "low": PatternFill("solid", fgColor="E8F5E9"),
    }
    review_fill = PatternFill("solid", fgColor="FFE0B2")

    for row, f in enumerate(findings, 2):
        conf = getattr(f, "confidence", 0.0) or 0.0
        values = [
            f.display_name, f.resource_name or "", f.resource_group or "",
            f.subscription_id or "", f.resource_type or "",
            round(f.estimated_savings_monthly, 2), round(f.estimated_savings_annual, 2),
            f.severity.upper(), round(conf, 2), _confidence_label(conf),
            (getattr(f, "validation_status", None) or "unvalidated"),
            getattr(f, "validation_variance_pct", None),
            getattr(f, "actual_monthly_cost", None),
            getattr(f, "advisor_recommendation_id", None) or "",
            f.description or "", f.recommendation or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row, col, val)
            if col in (6, 7, 13):
                cell.number_format = '#,##0.00'
            if col == 8:
                cell.fill = sev_fills.get(f.severity.lower(), PatternFill())
            if col == 11 and (getattr(f, "validation_status", None) or "") == "needs_review":
                cell.fill = review_fill

    col_widths = [28, 24, 18, 34, 30, 20, 20, 12, 11, 14, 14, 11, 20, 30, 45, 45]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    # ── Validation sheet (needs-review detail) ─────────────────────────────
    ws_val = wb.create_sheet("Validation")
    val_headers = ["Category", "Resource Name", "Estimated Monthly (USD)",
                   "Actual Monthly (USD)", "Variance %", "Note"]
    for col, h in enumerate(val_headers, 1):
        cell = ws_val.cell(1, col, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r = 2
    for f in findings:
        if (getattr(f, "validation_status", None) or "") != "needs_review":
            continue
        note = (f.details or {}).get("validation_note", "") if getattr(f, "details", None) else ""
        ws_val.cell(r, 1, f.display_name)
        ws_val.cell(r, 2, f.resource_name or "")
        ws_val.cell(r, 3, round(f.estimated_savings_monthly, 2))
        ws_val.cell(r, 4, getattr(f, "actual_monthly_cost", None))
        ws_val.cell(r, 5, getattr(f, "validation_variance_pct", None))
        ws_val.cell(r, 6, note)
        r += 1
    for col, width in enumerate([28, 24, 22, 22, 12, 60], 1):
        ws_val.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws_val.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
